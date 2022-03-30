from openpyxl import load_workbook
import urllib.request
import os
ERROR_HTTPError = -401
ERROR_URLError  = -402


def mkdir(path):
    folder = os.path.exists(path)

    if not folder:
        os.makedirs(path)
    else:
        print("下载选择的文件夹已存在")


def ifDownloadFile(img_name):
    if os.path.isfile(img_name):
        if os.path.getsize(img_name):
            return False
        else:
            fp = open(img_name, 'w')
            fp.close()
            return True
    else:
        fp = open(img_name, 'w')
        fp.close()
        return True


def download_img(img_url, img_name):
    try:
        req = urllib.request.Request(url=img_url, headers=kv)
        conn = urllib.request.urlopen(req)
        f = open(img_name, 'wb')
        f.write(conn.read())
    except urllib.error.HTTPError:
        print('\r\n%d Pic Saved Fail!' % (i-1))
        os.remove(img_name)
        return ERROR_HTTPError
    except urllib.error.URLError:
        print('\r\n%d Pic Saved Fail!' % (i-1))
        os.remove(img_name)
        return ERROR_URLError
    except:
        print('\r\n%d Pic Saved Fail!' % (i-1))
        f.close()
        os.remove(img_name)
    else:
        f.close()
        print('\r{:0>3} Saved Succ!'.format(i-1), end='')
        return 0


if __name__ == '__main__':
    # file = '2'
    # wb = load_workbook('C:/Users/Administrator/Desktop/' + file + '.xlsx')
    file = input("\n文件路径：\n")
    wb = load_workbook(file.replace("\"", ""))
    sheetnames = wb.sheetnames
    ws = wb[sheetnames[0]]  # index为0为第一张表
    print(ws.title)
    print(ws.max_row)
    print(ws.max_column)
    kv = {"user-agent": "Mozilla/5.0"}
    # path = "C:\\Users\\Administrator\\Desktop\\" + file + '\\'
    path = file.replace("\"", "").split('.')[0] + '\\'
    print(path)
    mkdir(path)
    # '''
    print("正在开始下载...")
    for i in range(2, ws.max_row + 1):
        name = path + str(i-1) + '.jpg'
        if ifDownloadFile(name):
            url = ws.cell(i, 2).value
            downResult = download_img(url, name)
            if downResult != 0:
                if downResult == ERROR_HTTPError:
                    print("图片链接异常,请检查表格链接")
                    break
                elif downResult == ERROR_URLError and i == 2:
                    print("请确保外网可用")
                    break
                elif downResult == ERROR_URLError and "链接无法找到" in ws.cell(i, 2).value:
                    print("该链接无效")
        else:
            print(f"\r{i-1}.jpg Exist", end='')
    # '''
